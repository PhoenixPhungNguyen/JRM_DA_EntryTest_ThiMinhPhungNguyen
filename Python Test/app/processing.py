from openpyxl import load_workbook
import math
import logging 
logger = logging.getLogger(__name__)

class ProcessPartA:
    def __init__(self):
        """
        Initialize the ProcessPartA instance with default data.
        """
        self.data = {
            "drillholes": [],
            "samples": [],
            "extra": [],
        }
        
    def retrieve_data(self, file_name):
        """
        Load data from an Excel workbook and run all analysis methods.
        Args:
            file_name (str): Path to the Excel file.
        """
        try:
            workbook = load_workbook(filename=file_name)
        except FileNotFoundError:
            logger.error(f"File not found: {file_name}")
            return
        except Exception as e:
            logger.exception(f"Failed to load workbook: {e}")
            return
        
        # Load DRILLHOLES
        try:
            worksheet = workbook["DRILLHOLES"]
            headers = [cell.value for cell in worksheet[1]]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                self.data["drillholes"].append(dict(zip(headers, row)))
        except KeyError:
            logger.error("Sheet 'DRILLHOLES' does not exist in the Excel file.")
        except Exception as e:
            logger.exception(f"Error occurred while reading 'DRILLHOLES' sheet: {e}")

        # Load SAMPLES
        try:
            worksheet = workbook["SAMPLES"]
            headers = [cell.value for cell in worksheet[1]]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                self.data["samples"].append(dict(zip(headers, row)))
        except KeyError:
            logger.error("Sheet 'SAMPLES' does not exist in the Excel file.")
        except Exception as e:
            logger.exception(f"Error occurred while reading 'SAMPLES' sheet: {e}")

        # Load EXTRA_DH_DATA
        try:
            worksheet = workbook["EXTRA_DH_DATA"]
            headers = [cell.value for cell in worksheet[1]]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                self.data["extra"].append(dict(zip(headers, row)))
        except KeyError:
            logger.error("Sheet 'EXTRA_DH_DATA' does not exist in the Excel file.")
        except Exception as e:
            logger.exception(f"Error occurred while reading 'EXTRA_DH_DATA' sheet: {e}")

        # Analyze everything
        try:
            logger.info("Analyzing data...")
            self.total_drill_length()
            self.avg_aud_grade()
            self.total_drill_meter_by_com()
            self.sum_drill_meter_by_com()
            self.total_drill_meter_by_day()
            logger.info("=== Distances from New Hole (2.5, 32.5) ===")
            self.distance_from_new_hole(2.5, 32.5)
            logger.info("=== Estimate Au Grade for New Hole (2.5, 32.5) ===")
            self.estimate_au_grade_for_new_hole(2.5, 32.5)
            logger.info("=== Estimate Au Grade for New Hole (7.5, 32.5) ===")
            self.estimate_au_grade_for_new_hole(7.5, 32.5)
            logger.info("=== Estimate Au Grade for New Hole (2.5, 37.5) ===")
            self.estimate_au_grade_for_new_hole(2.5, 37.5)
            logger.info("=== Estimate Au Grade for New Hole (7.5, 37.5) ===")
            self.estimate_au_grade_for_new_hole(7.5, 37.5)
        except Exception as e:
            logger.exception(f"Error during analysis: {e}")

    #Total drilled length
    def total_drill_length(self):
        """
        Calculate and log the total drilled length from drillholes.
        """
        drilled = 0
        for hole in self.data["drillholes"]:
            try:
                drilled += hole["Length (m)"]
            except KeyError:
                logger.warning(f"Missing 'Length (m)' in hole data: {hole}")
        logger.info(f"Total drilled length: {drilled} meters")

    #Average AU Grade
    def avg_aud_grade(self):
        """
        Calculate and log the average Au grade
        """
        total_grade = 0
        count = 0
        for sample in self.data["samples"]:
            try:
                total_grade += sample["Au"]
                count += 1
            except KeyError:
                logger.warning(f"Missing 'Au' in sample: {sample}")
            except TypeError:
                logger.warning(f"Invalid Au value in sample: {sample}")
        average_grade = 0.0
        if count>0:
            average_grade = round((total_grade / count),3)
        else:
            average_grade = 0.0
        logger.info(f"Average Au grade: {average_grade}")
        return average_grade

    #Report on the Total Drilled Meters by Company
    def total_drill_meter_by_com(self):
        """
        Build a mapping of drillhole names to drilling companies.
        """
        # Create a mapping of drillhole names to companies
        self.company_map = {}
        for item in self.data["extra"]:
            try:
                if item["Item"] == "COMPANY":
                    self.company_map[item["Name"]] = item["Value"]   #MRM001= DESRIG,MRM002= DESRIG,MRM003=HOLESRUS..
            except KeyError:
                logger.warning(f"Missing 'Item', 'Name', or 'Value' in extra data: {item}")

     ##Sum of the drilled meters by company
    def sum_drill_meter_by_com(self):
        """
        Calculate and log the total drilled meters by each company.
        """
        company_drill_lengths = {}
        for hole in self.data["drillholes"]:
            try:
                name = hole["Name"]         #MRM001
                length = hole["Length (m)"] #20
                company = self.company_map.get(name)       
                #check if the company is not in the mapping
                if company not in company_drill_lengths:
                    company_drill_lengths[company] = 0.0
                #add the length to the company drill lengths
                company_drill_lengths[company] += length
            except KeyError:
                logger.warning(f"Missing 'Name' or 'Length (m)' in hole: {hole}")
        logger.info("=== Report on the Total Drilled Meters by Company  ===")
        for company, total_length in company_drill_lengths.items():
            logger.info(f"- {company}: {total_length} meters")

    # Report on the Total Meters Drilled Per Day (1st, 2nd, 3rd, and 4th October) 
    def total_drill_meter_by_day(self):
        """
        Calculate and log the total meters drilled for each specific day.
        """
        # Create a mapping of drillhole names to dates
        date_map = {}
        for item in self.data["extra"]:
            try:
                if item["Item"] == "DATE":
                    date_map[item["Name"]] = item["Value"] #MRM001 = 01-10-2016
            except KeyError:
                logger.warning(f"Missing DATE information in item: {item}")
        # Sum the drilled meters per day
        daily_drill_lengths = {
            "01-10-2016": 0.0,
            "02-10-2016": 0.0,
            "03-10-2016": 0.0,
            "04-10-2016": 0.0,
        }
        for hole in self.data["drillholes"]:
            try:
                name = hole["Name"] #MRM001
                length = hole["Length (m)"] #20
                drill_date = date_map.get(name)
                if drill_date in daily_drill_lengths:
                    daily_drill_lengths[drill_date] += length
            except KeyError:
                logger.warning(f"Missing data for hole: {hole}")
        logger.info("=== Report on the Total Meters Drilled Per Day ===")
        for date in ["01-10-2016", "02-10-2016", "03-10-2016", "04-10-2016"]:
            logger.info(f"- {date}: {daily_drill_lengths[date]} meters")
    
    #Calculate Distances from New Hole
    def distance_from_new_hole(self,new_x, new_y):
        """
        Calculate and log distances from a given coordinate to all drillholes.

        Args:
            new_x (float): X coordinate of the new hole.
            new_y (float): Y coordinate of the new hole.
        """
        for hole in self.data["drillholes"]:
            try:
                name = hole["Name"]
                x = hole["X"]
                y = hole["Y"]
                distance = ((x - new_x) ** 2 + (y - new_y) ** 2) ** 0.5
                logger.info(f"- {name}: {round(distance, 3)} meters")
            except KeyError:
                logger.warning(f"Missing coordinates or name in hole: {hole}")
            except TypeError:
                logger.warning(f"Invalid coordinate data in hole: {hole}")

    # Estimate Au Grade for New Hole)
    def estimate_au_grade_for_new_hole(self, new_x, new_y):
        """
        Estimate and log the average Au grade based on the 4 closest drillholes.

        Args:
            new_x (float): X coordinate of the new hole.
            new_y (float): Y coordinate of the new hole.
        """

        drillhole_averages = []

        # Calculate distance of all drillholes to new point
        for hole in self.data["drillholes"]:
            try:
                name = hole["Name"]
                x, y = hole["X"], hole["Y"]
                distance = math.sqrt((x - new_x) ** 2 + (y - new_y) ** 2)

                # Find all samples belong to this drillholes
                samples = [s["Au"] for s in self.data["samples"] if s["Name"] == name and isinstance(s["Au"], (int, float))]
                
                if samples:
                    avg_au = sum(samples) / len(samples)
                    drillhole_averages.append((distance, avg_au))
                else:
                    logger.warning(f"No valid samples for drillhole {name}")
            
            except KeyError as e:
                logger.warning(f"Missing key {e} in drillhole or sample data")
            except TypeError:
                logger.warning(f"Invalid type in drillhole: {hole}")

        # Get closet drillholes
        closest_4 = sorted(drillhole_averages, key=lambda x: x[0])[:4]

        if closest_4:
            # Calculate Au Average from 4 average value of  drillholes
            average_au_grade = sum(avg for _, avg in closest_4) / len(closest_4)
            average_au_grade = round(average_au_grade, 3)
        else:
            average_au_grade = 0.0
            logger.warning("No nearby drillholes found within range")

        logger.info(f"Estimated Average Au Grade for New Hole ({new_x}, {new_y}): {average_au_grade}")
